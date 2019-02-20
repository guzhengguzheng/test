import openpyxl
from my_work.common import contants
#总思想 有两个类 对象（属性） 与excel相互赋值  最终目的 请求时 需要传入获取的数据去请求 也需要用到获取的数据来做预期
class Case:
     def __int__(self):   #要把属性值定义为初始化函数 这样每次调用对象都会默认执行init
           self.id=None   # 即每次调用对象属性值都是None 方便赋值
           self.title =None
           self.url =None
           self.data =None
           self.method =None
           self.expected =None
           self.actual=None
           self.result=None

class DoExcel:
    def __init__(self,filename,sheetname):
        self.wb=openpyxl.load_workbook(filename) #打开工作簿 open函数跟read函数都需要接上文件名（带路径）
        self.sheet=self.wb[sheetname]    #定位到具体某一个表单

    def read(self):
        max_row=self.sheet.max_row  #固定max_row读取所取表单最大行（表单的行）
        cases=[]   #建立空列表
        for r in range(2,max_row+1):  #赋值  excel赋值给对象的属性 方便对象在执行case的时候输入数据
            case = Case()   #创建对象 一定要放在这里 每循环一次 就创建一个新对象
            case.id=self.sheet.cell(row=r,column=1).value  #表单的单元格
            case.title=self.sheet.cell(row=r,column=2).value
            case.url=self.sheet.cell(row=r,column=3).value
            case.data=self.sheet.cell(row=r,column=4).value
            case.method=self.sheet.cell(row=r,column=5).value
            case.expected=self.sheet.cell(row=r,column=6).value
            cases.append(case)
        return cases

    def write(self,actual,result,r):  #赋值 执行完case 方便对象的属性 赋值给excel
        actual=self.sheet.cell(row=r,column=7).value
        result = self.sheet.cell(row=r, column=8).value
        self.wb.close()   #打开工作簿 就要有关闭工作簿的操作


if __name__ == '__main__':
    excel=DoExcel(contants.excel_dir,'login')
    cases=excel.read()
    print(cases)

#主要思路
# 创建对象来赋值
#对象的属性值 跟excel的单元格互相赋值
#从excel中读数据赋值给对象去执行case 执行case需要用到（url，method，data，expected）
#执行完case 输出（actual，result）赋给excel，写入结果